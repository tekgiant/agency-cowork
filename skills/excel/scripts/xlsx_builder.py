"""
Excel Workbook Builder.

Creates .xlsx files from structured content using openpyxl.
Reads branding and formatting settings from agentconfig.json.

Usage:
    python -m scripts.xlsx_builder --json data.json --output workbook.xlsx
    python -m scripts.xlsx_builder --sheets-json '[{"name":"Sheet1","headers":[...],"rows":[...]}]'
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from office_common.config import load_config, OfficeConfig


def hex_to_fill(hex_color: str) -> str:
    """Convert '#RRGGBB' to openpyxl fill color string (no hash)."""
    return hex_color.lstrip("#")


class XlsxBuilder:
    """Build Excel workbooks from structured content."""

    def __init__(self, config: Optional[OfficeConfig] = None):
        self.config = config or load_config()
        self.brand = self.config.branding
        self.xl_cfg = self.config.excel
        self.wb = Workbook()
        # Remove default sheet (we'll create named sheets)
        self._default_removed = False

    def _remove_default_sheet(self):
        if not self._default_removed and "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
            self._default_removed = True

    def _header_font(self) -> Font:
        return Font(
            name=self.brand.font_heading,
            size=self.xl_cfg.default_font_size_pt,
            bold=True,
            color=hex_to_fill(self.xl_cfg.header_font_color),
        )

    def _header_fill(self) -> PatternFill:
        return PatternFill(
            start_color=hex_to_fill(self.xl_cfg.header_fill_color),
            end_color=hex_to_fill(self.xl_cfg.header_fill_color),
            fill_type="solid",
        )

    def _stripe_fill(self) -> PatternFill:
        return PatternFill(
            start_color=hex_to_fill(self.xl_cfg.stripe_color),
            end_color=hex_to_fill(self.xl_cfg.stripe_color),
            fill_type="solid",
        )

    def _body_font(self) -> Font:
        return Font(
            name=self.xl_cfg.default_font,
            size=self.xl_cfg.default_font_size_pt,
        )

    def _thin_border(self) -> Border:
        side = Side(style="thin", color="D0D0D0")
        return Border(top=side, bottom=side, left=side, right=side)

    def add_sheet(self, name: str, headers: list[str], rows: list[list],
                  column_widths: Optional[list[int]] = None,
                  number_formats: Optional[dict[int, str]] = None,
                  formulas: Optional[list[dict]] = None):
        """Add a worksheet with headers, data, and formatting.

        Args:
            name: Sheet tab name
            headers: Column header labels
            rows: 2D list of cell values
            column_widths: Optional per-column widths (characters)
            number_formats: Optional dict of {col_index: format_string}
            formulas: Optional list of {"cell": "A10", "formula": "=SUM(A2:A9)"}
        """
        self._remove_default_sheet()
        ws = self.wb.create_sheet(title=name)

        header_font = self._header_font()
        header_fill = self._header_fill()
        stripe_fill = self._stripe_fill()
        body_font = self._body_font()
        border = self._thin_border()

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # Apply number format
                if number_formats and (col_idx - 1) in number_formats:
                    cell.number_format = number_formats[col_idx - 1]

                # Alternate row shading
                if row_idx % 2 == 0:
                    cell.fill = stripe_fill

        # Apply formulas
        if formulas:
            for f in formulas:
                ws[f["cell"]] = f["formula"]
                ws[f["cell"]].font = body_font
                ws[f["cell"]].border = border

        # Auto-fit column widths
        if column_widths:
            for i, w in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i + 1)].width = w
        else:
            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row in rows:
                    if col_idx - 1 < len(row):
                        max_len = max(max_len, len(str(row[col_idx - 1])))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Auto filter
        if self.xl_cfg.auto_filter and headers:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

        # Freeze header row
        if self.xl_cfg.freeze_header_row:
            ws.freeze_panes = "A2"

        return ws

    def add_chart(self, sheet_name: str, chart_type: str, title: str,
                  data_range: str, categories_range: Optional[str] = None,
                  position: str = "E2", style: int = 10):
        """Add a chart to an existing sheet.

        Args:
            sheet_name: Target worksheet name
            chart_type: "bar", "line", or "pie"
            title: Chart title
            data_range: Data reference (e.g., "B1:B10")
            categories_range: Category labels (e.g., "A2:A10")
            position: Cell to anchor the chart (e.g., "E2")
            style: Chart style number
        """
        ws = self.wb[sheet_name]

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.title = title
        chart.style = style

        # Parse range strings
        data_ref = Reference(ws, range_string=f"'{sheet_name}'!{data_range}")
        chart.add_data(data_ref, titles_from_data=True)

        if categories_range:
            cats = Reference(ws, range_string=f"'{sheet_name}'!{categories_range}")
            chart.set_categories(cats)

        ws.add_chart(chart, position)
        return chart

    def build_from_json(self, sheets_data: list[dict]):
        """Build workbook from a list of sheet definitions.

        Each sheet dict has:
            name: Sheet tab name
            headers: list of column headers
            rows: list of row data (list of lists)
            column_widths: (optional) list of widths
            number_formats: (optional) dict of {col_index: format}
            formulas: (optional) list of {"cell": "A10", "formula": "=SUM(...)"}
            charts: (optional) list of chart definitions
        """
        for sd in sheets_data:
            ws = self.add_sheet(
                name=sd.get("name", "Sheet1"),
                headers=sd.get("headers", []),
                rows=sd.get("rows", []),
                column_widths=sd.get("column_widths"),
                number_formats=sd.get("number_formats"),
                formulas=sd.get("formulas"),
            )

            for chart in sd.get("charts", []):
                self.add_chart(
                    sheet_name=sd["name"],
                    chart_type=chart.get("type", "bar"),
                    title=chart.get("title", ""),
                    data_range=chart.get("data_range", ""),
                    categories_range=chart.get("categories_range"),
                    position=chart.get("position", "E2"),
                )

    def save(self, output_path: Optional[str] = None) -> str:
        """Save the workbook and return the path."""
        if output_path is None:
            os.makedirs(self.config.output_dir, exist_ok=True)
            output_path = os.path.join(self.config.output_dir, "workbook.xlsx")
        else:
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        self.wb.save(output_path)
        return output_path


def main():
    parser = argparse.ArgumentParser(description="Build Excel workbooks")
    parser.add_argument("--json", help="Path to JSON file with sheet definitions")
    parser.add_argument("--sheets-json", help="Inline JSON string with sheet definitions")
    parser.add_argument("--output", "-o", help="Output .xlsx file path")
    parser.add_argument("--config", help="Path to agentconfig.json")
    args = parser.parse_args()

    config = load_config(args.config)
    builder = XlsxBuilder(config)

    if args.json:
        with open(args.json, "r", encoding="utf-8") as f:
            sheets = json.load(f)
        builder.build_from_json(sheets)
    elif args.sheets_json:
        sheets = json.loads(args.sheets_json)
        builder.build_from_json(sheets)
    else:
        data = json.load(sys.stdin)
        builder.build_from_json(data)

    path = builder.save(args.output)
    sheets_info = [{"name": s, "rows": ws.max_row - 1, "cols": ws.max_column}
                   for s, ws in zip(builder.wb.sheetnames, builder.wb.worksheets)]
    print(json.dumps({"status": "ok", "path": os.path.abspath(path), "sheets": sheets_info}))


if __name__ == "__main__":
    main()
