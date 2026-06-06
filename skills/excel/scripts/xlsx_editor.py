"""
Excel Workbook Editor.

Primary tool for inspecting and modifying existing .xlsx files.
Workflow: download from SharePoint → inspect → edit → upload back.

Usage:
    python -m scripts.xlsx_editor --input workbook.xlsx --action inspect
    python -m scripts.xlsx_editor --input workbook.xlsx --action inspect --sheet "Sheet1"
    python -m scripts.xlsx_editor --input workbook.xlsx --action extract-data --sheet "Sheet1"
    python -m scripts.xlsx_editor --input workbook.xlsx --action find-replace --find "OLD" --replace "NEW"
    python -m scripts.xlsx_editor --input workbook.xlsx --action update-cells --sheet "S1" --updates-json '[...]'
    python -m scripts.xlsx_editor --input workbook.xlsx --action add-rows --sheet "S1" --rows-json '[...]'
    python -m scripts.xlsx_editor --input workbook.xlsx --action insert-rows --sheet "S1" --at-row 5 --rows-json '[...]'
    python -m scripts.xlsx_editor --input workbook.xlsx --action delete-rows --sheet "S1" --rows "5,6,7"
    python -m scripts.xlsx_editor --input workbook.xlsx --action add-sheet --sheets-json '[...]'
    python -m scripts.xlsx_editor --input workbook.xlsx --action delete-sheet --sheet "Sheet2"
    python -m scripts.xlsx_editor --input workbook.xlsx --action copy-sheet --sheet "S1" --new-name "S1 Copy"
    python -m scripts.xlsx_editor --input workbook.xlsx --action rename-sheet --sheet "S1" --new-name "Summary"
    python -m scripts.xlsx_editor --input workbook.xlsx --action batch --ops-json '[...]'
"""

import argparse
import copy
import json
import os
import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from office_common.config import load_config


# ─── Inspect ──────────────────────────────────────────────────────────────────

def inspect_workbook(wb, sheet_name: Optional[str] = None, max_preview_rows: int = 5) -> dict:
    """Comprehensive inspection of a workbook or single sheet.

    Returns structure with enough detail for the agent to plan edits:
    - Sheet names, dimensions, merged cells, filters
    - Header row + sample data rows
    - Formula locations
    - Chart info
    """
    info = {
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
    }

    sheets_to_inspect = []
    if sheet_name:
        if sheet_name in wb.sheetnames:
            sheets_to_inspect = [(sheet_name, wb[sheet_name])]
        else:
            info["error"] = f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            return info
    else:
        sheets_to_inspect = [(n, wb[n]) for n in wb.sheetnames]

    info["sheets"] = []
    for name, ws in sheets_to_inspect:
        sheet_info = _inspect_sheet(name, ws, max_preview_rows)
        info["sheets"].append(sheet_info)

    return info


def _inspect_sheet(name: str, ws, max_preview_rows: int) -> dict:
    """Inspect a single sheet."""
    sheet = {
        "name": name,
        "dimensions": ws.dimensions,
        "rows": ws.max_row,
        "columns": ws.max_column,
        "merged_cells": [str(m) for m in ws.merged_cells.ranges] if ws.merged_cells else [],
        "frozen_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
    }

    # Headers (row 1)
    headers = []
    if ws.max_row >= 1:
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row=1, column=col)
            headers.append(str(cell.value) if cell.value is not None else "")
    sheet["headers"] = headers

    # Column widths
    col_widths = {}
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width and dim.width != 8.43:  # Skip default
            col_widths[col_letter] = round(dim.width, 1)
    if col_widths:
        sheet["column_widths"] = col_widths

    # Sample data (first N rows after header)
    preview_rows = []
    for row_idx in range(2, min(ws.max_row + 1, max_preview_rows + 2)):
        row_data = []
        for col in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(row=row_idx, column=col)
            row_data.append(cell.value if cell.value is not None else "")
        preview_rows.append({"row": row_idx, "data": row_data})
    sheet["preview_rows"] = preview_rows

    # Last few rows (for seeing the end of the data)
    if ws.max_row > max_preview_rows + 2:
        tail_rows = []
        start = max(ws.max_row - 2, max_preview_rows + 2)
        for row_idx in range(start, ws.max_row + 1):
            row_data = []
            for col in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=row_idx, column=col)
                row_data.append(cell.value if cell.value is not None else "")
            tail_rows.append({"row": row_idx, "data": row_data})
        sheet["tail_rows"] = tail_rows

    # Formulas (scan for cells with formulas)
    formulas = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({
                    "cell": cell.coordinate,
                    "formula": cell.value,
                })
    if formulas:
        sheet["formulas"] = formulas[:50]  # Cap at 50

    # Charts
    if ws._charts:
        sheet["charts"] = [{"title": c.title, "type": type(c).__name__}
                           for c in ws._charts]

    return sheet


# ─── Edit Operations ──────────────────────────────────────────────────────────

def find_replace(wb, find: str, replace: str,
                 sheet_name: Optional[str] = None) -> int:
    """Find and replace text across sheets. Returns replacement count."""
    count = 0
    sheets = [wb[sheet_name]] if sheet_name else [wb[n] for n in wb.sheetnames]
    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and find in cell.value:
                    cell.value = cell.value.replace(find, replace)
                    count += 1
    return count


def update_cells(wb, sheet_name: str, updates: list[dict]) -> int:
    """Update specific cells.

    Each update: {"cell": "A1", "value": "x"} or {"row": 2, "col": 1, "value": "x"}
    """
    ws = wb[sheet_name]
    count = 0
    for u in updates:
        if "cell" in u:
            ws[u["cell"]] = u["value"]
            count += 1
        elif "row" in u and "col" in u:
            ws.cell(row=u["row"], column=u["col"], value=u["value"])
            count += 1
    return count


def add_rows(wb, sheet_name: str, rows: list[list]) -> int:
    """Append rows at the end of a sheet."""
    ws = wb[sheet_name]
    for row in rows:
        ws.append(row)
    return len(rows)


def insert_rows(wb, sheet_name: str, at_row: int, rows: list[list]) -> int:
    """Insert rows at a specific position (shifts existing rows down)."""
    ws = wb[sheet_name]
    ws.insert_rows(at_row, amount=len(rows))
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            ws.cell(row=at_row + i, column=j + 1, value=val)
    return len(rows)


def delete_rows(wb, sheet_name: str, row_indices: list[int]) -> int:
    """Delete rows by index (1-based). Processes in reverse to maintain indices."""
    ws = wb[sheet_name]
    count = 0
    for idx in sorted(row_indices, reverse=True):
        ws.delete_rows(idx, 1)
        count += 1
    return count


def delete_sheet(wb, sheet_name: str):
    """Delete a sheet by name."""
    del wb[sheet_name]


def copy_sheet(wb, sheet_name: str, new_name: str):
    """Copy a sheet within the workbook."""
    source = wb[sheet_name]
    target = wb.copy_worksheet(source)
    target.title = new_name
    return new_name


def rename_sheet(wb, sheet_name: str, new_name: str):
    """Rename a sheet."""
    wb[sheet_name].title = new_name


# ─── Batch Operations ─────────────────────────────────────────────────────────

def batch_ops(wb, operations: list[dict]) -> list[dict]:
    """Execute a batch of edit operations.

    Supported actions:
        find-replace: {find, replace, sheet (optional)}
        update-cells: {sheet, updates: [{cell, value}]}
        add-rows: {sheet, rows: [[...]]}
        insert-rows: {sheet, at_row, rows: [[...]]}
        delete-rows: {sheet, rows: [5,6,7]}
        delete-sheet: {sheet}
        copy-sheet: {sheet, new_name}
        rename-sheet: {sheet, new_name}
    """
    results = []
    for op in operations:
        action = op.get("action")
        try:
            if action == "find-replace":
                count = find_replace(wb, op["find"], op["replace"], op.get("sheet"))
                results.append({"action": action, "ok": True, "replacements": count})
            elif action == "update-cells":
                count = update_cells(wb, op["sheet"], op["updates"])
                results.append({"action": action, "ok": True, "updates": count})
            elif action == "add-rows":
                count = add_rows(wb, op["sheet"], op["rows"])
                results.append({"action": action, "ok": True, "added": count})
            elif action == "insert-rows":
                count = insert_rows(wb, op["sheet"], op["at_row"], op["rows"])
                results.append({"action": action, "ok": True, "inserted": count})
            elif action == "delete-rows":
                count = delete_rows(wb, op["sheet"], op["rows"])
                results.append({"action": action, "ok": True, "deleted": count})
            elif action == "delete-sheet":
                delete_sheet(wb, op["sheet"])
                results.append({"action": action, "ok": True})
            elif action == "copy-sheet":
                name = copy_sheet(wb, op["sheet"], op["new_name"])
                results.append({"action": action, "ok": True, "new_name": name})
            elif action == "rename-sheet":
                rename_sheet(wb, op["sheet"], op["new_name"])
                results.append({"action": action, "ok": True})
            else:
                results.append({"action": action, "ok": False, "error": f"Unknown action: {action}"})
        except Exception as e:
            results.append({"action": action, "ok": False, "error": str(e)})
    return results


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Inspect and edit Excel workbooks")
    parser.add_argument("--input", "-i", required=True, help="Input .xlsx file")
    parser.add_argument("--output", "-o", help="Output file (default: overwrites input)")
    parser.add_argument("--action", required=True,
                        choices=["inspect", "extract-data", "find-replace",
                                 "update-cells", "add-rows", "insert-rows",
                                 "delete-rows", "add-sheet", "delete-sheet",
                                 "copy-sheet", "rename-sheet", "batch"],
                        help="Action to perform")
    parser.add_argument("--sheet", help="Sheet name")
    parser.add_argument("--new-name", help="New sheet name (for copy/rename)")
    parser.add_argument("--find", help="Text to find")
    parser.add_argument("--replace", help="Replacement text")
    parser.add_argument("--at-row", type=int, help="Row position for insert")
    parser.add_argument("--rows", help="Comma-separated row indices for delete")
    parser.add_argument("--max-rows", type=int, default=1000, help="Max rows to extract")
    parser.add_argument("--preview-rows", type=int, default=5, help="Preview rows for inspect")
    # JSON args
    parser.add_argument("--sheets-json", help="JSON sheet definitions (for add-sheet)")
    parser.add_argument("--updates-json", help="JSON cell updates")
    parser.add_argument("--rows-json", help="JSON rows to add/insert")
    parser.add_argument("--ops-json", help="JSON batch operations")
    parser.add_argument("--config", help="Path to agentconfig.json")
    args = parser.parse_args()

    wb = load_workbook(args.input)
    output_path = args.output or args.input

    if args.action == "inspect":
        result = inspect_workbook(wb, args.sheet, args.preview_rows)
        print(json.dumps(result, indent=2, default=str))
        return

    elif args.action == "extract-data":
        if not args.sheet:
            args.sheet = wb.sheetnames[0]
        ws = wb[args.sheet]
        headers = []
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [str(c) if c is not None else "" for c in row]
            elif row_idx <= args.max_rows + 1:
                rows.append([c if c is not None else "" for c in row])
        result = {"name": args.sheet, "headers": headers, "rows": rows,
                  "total_rows": ws.max_row - 1}
        print(json.dumps(result, indent=2, default=str))
        return

    elif args.action == "find-replace":
        if not args.find:
            print("Error: --find required", file=sys.stderr)
            sys.exit(1)
        count = find_replace(wb, args.find, args.replace or "", args.sheet)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "replacements": count}))

    elif args.action == "update-cells":
        if not args.sheet or not args.updates_json:
            print("Error: --sheet and --updates-json required", file=sys.stderr)
            sys.exit(1)
        updates = json.loads(args.updates_json)
        count = update_cells(wb, args.sheet, updates)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "updates": count}))

    elif args.action == "add-rows":
        if not args.sheet or not args.rows_json:
            print("Error: --sheet and --rows-json required", file=sys.stderr)
            sys.exit(1)
        rows = json.loads(args.rows_json)
        count = add_rows(wb, args.sheet, rows)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "added": count}))

    elif args.action == "insert-rows":
        if not args.sheet or not args.rows_json or args.at_row is None:
            print("Error: --sheet, --rows-json, --at-row required", file=sys.stderr)
            sys.exit(1)
        rows = json.loads(args.rows_json)
        count = insert_rows(wb, args.sheet, args.at_row, rows)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "inserted": count}))

    elif args.action == "delete-rows":
        if not args.sheet or not args.rows:
            print("Error: --sheet and --rows required", file=sys.stderr)
            sys.exit(1)
        row_indices = [int(x.strip()) for x in args.rows.split(",")]
        count = delete_rows(wb, args.sheet, row_indices)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "deleted": count}))

    elif args.action == "add-sheet":
        if not args.sheets_json:
            print("Error: --sheets-json required", file=sys.stderr)
            sys.exit(1)
        from scripts.xlsx_builder import XlsxBuilder
        config = load_config(args.config)
        builder = XlsxBuilder(config)
        builder.wb = wb
        builder._default_removed = True
        sheets = json.loads(args.sheets_json)
        builder.build_from_json(sheets)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "total_sheets": len(wb.sheetnames)}))

    elif args.action == "delete-sheet":
        if not args.sheet:
            print("Error: --sheet required", file=sys.stderr)
            sys.exit(1)
        delete_sheet(wb, args.sheet)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "remaining": wb.sheetnames}))

    elif args.action == "copy-sheet":
        if not args.sheet or not args.new_name:
            print("Error: --sheet and --new-name required", file=sys.stderr)
            sys.exit(1)
        name = copy_sheet(wb, args.sheet, args.new_name)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "new_name": name}))

    elif args.action == "rename-sheet":
        if not args.sheet or not args.new_name:
            print("Error: --sheet and --new-name required", file=sys.stderr)
            sys.exit(1)
        rename_sheet(wb, args.sheet, args.new_name)
        wb.save(output_path)
        print(json.dumps({"status": "ok"}))

    elif args.action == "batch":
        if not args.ops_json:
            print("Error: --ops-json required", file=sys.stderr)
            sys.exit(1)
        ops = json.loads(args.ops_json)
        results = batch_ops(wb, ops)
        wb.save(output_path)
        print(json.dumps({"status": "ok", "results": results}))


if __name__ == "__main__":
    main()
